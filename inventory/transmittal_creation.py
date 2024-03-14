import os
from django.contrib import admin
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponseRedirect, HttpResponse
from django.shortcuts import get_object_or_404
from django.urls import reverse
from openpyxl import load_workbook
from copy import copy

from openpyxl import styles
from openpyxl.styles import Alignment
from inventory.models import CleanlinessLevel, Contaminant, Damage, Material, WorkOrder, Item, WorkOrderItem, BubblePointLog, \
    LabLog, LabLogItem, ItemType, Manufacturer, Status, Location, Fitting, LabLogFitting, WorkOrderFitting, FittingSize, \
    Customer

TEMPLATE_PATH = 'staticfiles/report_template/transmittal_template.xlsx'


def create_transmittal(work_order_id):
    transmittal_cell_regs = {
        'TRANSMITTAL_DATE_CELL': 'D13',
        'ATTENTION_COLUMN': 'D14',
        'RE_COLUMN': 'D15',
        'DATE_COLUMN': 'A',
        'QTY_COLUMN': 'B',
        'MARS_ID_COLUMN': 'C',
        'TAG_NAME_COLUMN': 'D',
        'DESCRIPTION_COLUMN': 'E',
        'MANUFACTURER_COLUMN': 'F',
        'MODEL_COLUMN': 'G',
        'LOT_NUMBER_COLUMN': 'H'
    }
    current_row = 20
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active  # Assuming you're working with the first sheet
    work_order = WorkOrder.objects.filter(pk=work_order_id).select_related('status', 'location', 'customer').first()
    # Fetch all related WorkOrderItems for this WorkOrder
    work_order_items = WorkOrderItem.objects.filter(work_order=work_order).select_related('item')
    work_order_fittings = WorkOrderFitting.objects.filter(work_order=work_order).select_related(
        'work_order_unique_fitting')

    ws[f'{transmittal_cell_regs["TRANSMITTAL_DATE_CELL"]}'] = 'TEST'

    report_path = f'work_order_{work_order.pk}_report.xlsx'


create_transmittal(1)
