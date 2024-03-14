import os
from django.contrib import admin
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponseRedirect, HttpResponse
from django.shortcuts import get_object_or_404
from django.urls import reverse
from openpyxl import load_workbook
from copy import copy

from openpyxl import styles
from openpyxl.styles import Alignment
from .models import CleanlinessLevel, Contaminant, Damage, Material, WorkOrder, Item, WorkOrderItem, BubblePointLog, \
    LabLog, LabLogItem, ItemType, Manufacturer, Status, Location, Fitting, LabLogFitting, WorkOrderFitting, FittingSize, \
    Customer


def copy_row_formatting(ws, source_row, target_row):
    for col in range(1, ws.max_column + 1):
        source_cell = ws.cell(row=source_row, column=col)
        target_cell = ws.cell(row=target_row, column=col)

        # Copy the style from source_cell to target_cell
        if source_cell.has_style:
            target_cell.alignment = Alignment(wrap_text=True)
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
            # target_cell.alignment = Alignment(vertical='center')
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


@staff_member_required
def download_work_order_report(request, work_order_id):
    # Ensure the work order exists and handle permissions
    work_order = get_object_or_404(WorkOrder, pk=work_order_id)

    # Assuming generate_work_order_excel_report is updated as previously described
    file_path = generate_work_order_excel_report(work_order_id)

    with open(file_path, 'rb') as excel_file:
        response = HttpResponse(excel_file.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="work_order_{work_order_id}_report.xlsx"'
        return response


def generate_work_order_excel_report(work_order_id):
    # Load the work order with related objects
    work_order = WorkOrder.objects.filter(pk=work_order_id).select_related('status', 'location', 'customer').first()
    # Fetch all related WorkOrderItems for this WorkOrder
    work_order_items = WorkOrderItem.objects.filter(work_order=work_order).select_related('item')
    work_order_fittings = WorkOrderFitting.objects.filter(work_order=work_order).select_related(
        'work_order_unique_fitting')

    template_path = 'staticfiles/report_template/work_order_template.xlsx'
    report_path = f'work_order_{work_order.pk}_report.xlsx'

    # Load the first template
    wb = load_workbook(template_path)
    ws = wb.active  # Assuming you're working with the first sheet

    # Fill out customer here
    ws['A2'] = f'{work_order.customer.customer_name}'
    ws['A3'] = f'{work_order.customer.customer_address}'

    # Fill out MARS ID Here
    ws['AQ2'] = f'MARS -  {work_order.pk}'
    ws['AS5'] = f'MARS -  {work_order.pk}'

    # Fill out dates here
    ws['AB4'] = work_order.date_created.strftime('%B, %d %Y')
    ws['AB5'] = work_order.need_by_date.strftime('%B, %d %Y %H:%M%p')

    # Begin to fill out items here the top portion of the work order is static in height.
    source_row = 18
    item_count = 0
    ws[f'A{source_row}'].alignment = Alignment(vertical='center')
    for enum, item in enumerate(work_order_items, start=0):
        row_number = source_row + enum
        temp_item = f'MARS ID: {item.item.mars_id} Item: {item.item.item_name} PN: {item.item.serial_number}'
        if row_number > 18:
            ws.insert_rows(row_number)
            copy_row_formatting(ws, source_row, row_number)
            ws.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=60)
        ws[f'A{row_number}'] = temp_item
        item_count += 1

    for enum, item in enumerate(work_order_fittings, start=0):
        row_number = source_row + item_count
        temp_item = f'QTY: {item.quantity} {item.work_order_unique_fitting}'
        if row_number > 18:
            ws.insert_rows(row_number)
            copy_row_formatting(ws, source_row, row_number)
            ws.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=60)
        ws[f'A{row_number}'] = temp_item
        item_count += 1

    ws.merge_cells(start_row=source_row + item_count, start_column=1, end_row=source_row + item_count, end_column=60)

    # Add header for processing data
    copy_row_formatting(ws, 16, source_row + item_count)
    ws[f'A{source_row + item_count}'] = 'PROCESSING DATA'
    item_count += 1

    PROCESSING_LIST = {'material_type': None, 'component_data': None, 'service_media': None,
                       'cleaning_specification': None, }
    PROCESSING_LIST = get_material_types(work_order_id)
    ws[f'A{24}'] = str(PROCESSING_LIST['material_type'])

    # Save the filled report to a new file
    wb.save(report_path)
    return report_path


def get_material_types(work_order_id):
    PROCESSING_LIST = {'material_type': set(), 'component_data': None, 'service_media': None,
                       'cleaning_specification': None}

    # Load the work order with related objects
    work_order = WorkOrder.objects.filter(pk=work_order_id).select_related('status', 'location', 'customer').first()

    # Fetch all related WorkOrderItems for this WorkOrder
    work_order_items = WorkOrderItem.objects.filter(work_order=work_order).select_related('item')

    # Iterate over the work order items and add their material types to the set
    for item in work_order_items:
        if item.item.item_material:
            PROCESSING_LIST['material_type'].add(item.item.item_material)
    return PROCESSING_LIST


class WorkOrderItemInline(admin.TabularInline):
    model = WorkOrderItem
    extra = 1
    search_fields = ['item']
    autocomplete_fields = ['item']
    # fields = ['item']
    # autocomplete_fields = ['item']  # Replace 'item' with the correct field name if different


class LabLogItemInline(admin.TabularInline):
    model = LabLogItem
    extra = 1

    # Define fields and other properties as needed


class LabLogFittingInline(admin.TabularInline):
    model = LabLogFitting
    extra = 1


class FittingAdmin(admin.ModelAdmin):
    search_fields = ['fitting_name', 'description', 'lot_number']


class WorkOrderFittingInLine(admin.TabularInline):
    model = WorkOrderFitting
    extra = 1
    search_fields = ['work_order_unique_fitting']
    autocomplete_fields = ['work_order_unique_fitting']


class LabLogAdmin(admin.ModelAdmin):
    list_display = ('lab_log_id', 'date_logged', 'cleanliness_level')  # Adjust fields as needed
    list_filter = ('cleanliness_level',)  # Adjust or add more filters as needed
    search_fields = ('cleanliness_level', 'date_logged')  # Adjust search fields as needed
    inlines = [LabLogItemInline, LabLogFittingInline]  # Include this only if you defined the LabLogItemInline


# Define the custom admin class for work order
class WorkOrderAdmin(admin.ModelAdmin):
    list_display = ('work_order_id', 'date_created', 'description', 'cleanliness_level', 'need_by_date')
    list_filter = ('cleanliness_level',)
    search_fields = ('description', 'comment')
    readonly_fields = ('work_order_id',)  # Make the work_order_id field read-only
    inlines = [WorkOrderItemInline, WorkOrderFittingInLine]
    change_form_template = 'admin/inventory/workorder/change_form.html'


class ItemAdmin(admin.ModelAdmin):
    list_display = ('mars_id', 'item_name', 'item_type', 'model_number', 'manufacturer', 'serial_number')
    search_fields = ['mars_id', 'item_name', 'description']  # You can adjust the fields as needed


class CustomerAdmin(admin.ModelAdmin):
    list_display = ['customer_name', 'point_of_contact', 'phone_number', 'customer_address']


@staff_member_required
def download_transmittal(request, work_order_id):
    # Ensure the work order exists and handle permissions
    work_order = get_object_or_404(WorkOrder, pk=work_order_id)

    # Generate the transmittal Excel file
    file_path = create_transmittal(work_order_id)

    with open(file_path, 'rb') as excel_file:
        response = HttpResponse(excel_file.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="transmittal_{work_order_id}.xlsx"'
        return response


def create_transmittal(work_order_id):
    TEMPLATE_PATH = 'staticfiles/report_template/transmittal_template.xlsx'

    transmittal_cell_regs = {
        'TRANSMITTAL_DATE_CELL': 'D13',
        'ATTENTION_COLUMN': 'D14',
        'RE_COLUMN': 'D15',
        'DATE_COLUMN': 'A',
        'QTY_COLUMN': 'B',
        'MARS_ID_COLUMN': 'C',
        'TAG_NAME_COLUMN': 'D',
        'DESCRIPTION_COLUMN': 'E',
        'MANUFACTURER_COLUMN': 'F',
        'MODEL_COLUMN': 'G',
        'LOT_NUMBER_COLUMN': 'H'
    }
    static_row = 20
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active  # Assuming you're working with the first sheet
    work_order = WorkOrder.objects.filter(pk=work_order_id).select_related('status', 'location', 'customer').first()
    # Fetch all related WorkOrderItems for this WorkOrder
    work_order_items = WorkOrderItem.objects.filter(work_order=work_order).select_related('item')
    work_order_fittings = WorkOrderFitting.objects.filter(work_order=work_order).select_related(
        'work_order_unique_fitting')

    ws[f'{transmittal_cell_regs["TRANSMITTAL_DATE_CELL"]}'] = work_order.date_created.strftime('%m-%d-%Y')
    ws[f'{transmittal_cell_regs["ATTENTION_COLUMN"]}'] = 'STEVE PEYTON'
    ws[f'{transmittal_cell_regs["RE_COLUMN"]}'] = str(work_order.cleanliness_level)

    for enum, item in enumerate(work_order_items):
        current_row = static_row + enum
        copy_row_formatting(ws, static_row, current_row)
        ws[f'{transmittal_cell_regs['DATE_COLUMN']}{current_row}'] = work_order.date_created.strftime('%m-%d-%Y')
        ws[f'{transmittal_cell_regs['QTY_COLUMN']}{current_row}'] = 1
        ws[f'{transmittal_cell_regs['MARS_ID_COLUMN']}{current_row}'] = item.item.mars_id
        ws[f'{transmittal_cell_regs['TAG_NAME_COLUMN']}{current_row}'] = 'MARS'
        ws[f'{transmittal_cell_regs['DESCRIPTION_COLUMN']}{current_row}'] = item.item.description

    report_path = f'Transmittal_{work_order.pk}.xlsx'
    wb.save(report_path)
    return report_path


# Register models with their respective admin classes
admin.site.register(CleanlinessLevel)
admin.site.register(Contaminant)
admin.site.register(Damage)
admin.site.register(Material)
admin.site.register(WorkOrder, WorkOrderAdmin)
admin.site.register(Item, ItemAdmin)
admin.site.register(ItemType)
admin.site.register(BubblePointLog)
admin.site.register(LabLog, LabLogAdmin)
admin.site.register(LabLogItem)
admin.site.register(Manufacturer)
admin.site.register(Status)
admin.site.register(Location)
admin.site.register(LabLogFitting)
admin.site.register(WorkOrderFitting)
admin.site.register(Fitting, FittingAdmin)
admin.site.register(FittingSize)
admin.site.register(Customer, CustomerAdmin)
